#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Apr 10 19:01:55 2019

@author: ashik
"""

import pandas as pd
import time
import os
import glob
import datetime
from datetime import timedelta
import scipy.stats
import math
#articleDF = pd.read_excel("data/ajb9b3.xlsx")

#time.strftime("%A %Y-%m-%d %H:%M:%S", time.localtime(df.iloc[1,6]/1000))

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):

    from openpyxl import load_workbook
    
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        writer.book = load_workbook(filename)
        
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass

    if startrow is None:
        startrow = 0

    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()
    
    
    
def calculateWeekData(articleDF, tempIndex):
    
    day = int(time.strftime("%d", time.localtime(articleDF.iloc[tempIndex,6]/1000)))
    weekDate = time.strftime('%A', time.localtime(articleDF.iloc[tempIndex,5]/1000))

    while(weekDate!='Monday' and tempIndex<len(articleDF)):
        weekDate = time.strftime('%A', time.localtime(articleDF.iloc[tempIndex,5]/1000))   
        tempIndex = tempIndex+1
    
    print(weekDate)
        
    weekDayDF = pd.DataFrame()
    '''
    if tempIndex<len(articleDF):
        day = int(time.strftime("%d", time.localtime(articleDF.iloc[tempIndex,6]/1000)))'''
    nextWeek = day+5
    nextDay = 'Monday'
    #while (day<nextWeek and tempIndex<len(articleDF)):
    while (nextDay!='Saturday' and tempIndex<len(articleDF)):
        duration = int(articleDF.iloc[tempIndex,9])
        if (duration>=10):
            weekDay = time.strftime('%A', time.localtime(articleDF.iloc[tempIndex,5]/1000))
            startTime = time.strftime("%H:%M:%S", time.localtime(articleDF.iloc[tempIndex,5]/1000))
            endTime = time.strftime("%H:%M:%S", time.localtime(articleDF.iloc[tempIndex,6]/1000))
            timeRange = startTime+"-"+endTime
            dpd = float(articleDF.iloc[tempIndex, 3])/duration
            startTime = articleDF.iloc[tempIndex, 5]
            endTime = articleDF.iloc[tempIndex, 6]
            weekDate = time.strftime("%Y-%m-%d", time.localtime(articleDF.iloc[tempIndex,6]/1000))
            weekDayDF = weekDayDF.append({'Day': weekDay, 'Time':timeRange, 'Octets/Duration':dpd, 'Date':weekDate, 'Start time': startTime, 'End time':endTime}, ignore_index = True)
        
        #print(day)
        
        #day = int(time.strftime("%d", time.localtime(articleDF.iloc[tempIndex,6]/1000)))
        nextDay = time.strftime("%A", time.localtime(articleDF.iloc[tempIndex,6]/1000))
        tempIndex = tempIndex+1
        
    if (len(weekDayDF)):
        weekDayDF = weekDayDF.sort_values(by=['Start time'])
        weekDayDF = weekDayDF.reset_index(drop=True)
    else:
        startTime = articleDF.iloc[tempIndex-1,5]/1000
        weekDayDF = weekDayDF.append({'Day': 0, 'Time':0, 'Octets/Duration':0, 'Date':0, 'Start time': startTime, 'End time':0}, ignore_index = True)


    return weekDayDF, tempIndex
    

def getWeek(articleDF):
    
    tempIndex = 0
    firstWeek = pd.DataFrame()
    firstWeek, tempIndex = calculateWeekData(articleDF, tempIndex)
    
    print("first")
    
    secondWeek = pd.DataFrame()
    if (tempIndex<len(articleDF)):
        print("goto 2nd")
        print(tempIndex)
        secondWeek, tempIndex = calculateWeekData(articleDF, tempIndex)
    else:
        startTime = articleDF.iloc[tempIndex-1,5]/1000
        secondWeek = secondWeek.append({'Day': 0, 'Time':0, 'Octets/Duration':0, 'Date':0, 'Start time': startTime, 'End time':0}, ignore_index = True)
    
    return firstWeek, secondWeek

def epochTimeCreate(weekDataDF):
    
    epochListDF = pd.DataFrame()
    
    firstDay = weekDataDF.iloc[0,4]
    day = int(time.strftime("%d", time.localtime(firstDay/1000)))
    month = int(time.strftime("%m", time.localtime(firstDay/1000)))
    year = int(time.strftime("%Y", time.localtime(firstDay/1000)))
    
    timeFormat = "%Y-%m-%d %H:%M:%S"
    t1 = datetime.datetime(year, month, day, 8, 00, 00)
    t2 = datetime.datetime(year, month, day, 17, 00, 00)
    
    for i in range (0,5):
        startTime = t1 + timedelta(days=i)
        endTime = t2 + timedelta(days=i)
        epochInitial = int(time.mktime(time.strptime(str(startTime), timeFormat)))
        epochFinal = int(time.mktime(time.strptime(str(endTime), timeFormat)))
        
        while epochInitial<=epochFinal:
            curTime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(epochInitial))
            epochListDF = epochListDF.append({'Epoch list':str(epochInitial), 'Real time':curTime}, ignore_index=True)
            epochInitial = epochInitial+10
            
    return epochListDF
            

def splitData(epochListDF, weekDataDF, delN):    
    
    index=0
    i=0
    finalSplitList = pd.DataFrame()
    sameIndex=0
    flag=1
    while(i<len(epochListDF) and index<len(weekDataDF)):
        epochtime = int(epochListDF['Epoch list'][i])
        epochtime2 = int((weekDataDF['Start time'][index])/1000)
        startTime = time.strftime("%A -- %H:%M:%S --", time.localtime(epochtime))
        endtime = epochtime+delN
        endTime = time.strftime("%H:%M:%S", (time.localtime(endtime)))
        timeRange = startTime + endTime
        
        if ((epochtime2>=epochtime and epochtime2<endtime) and (len(finalSplitList)!=0)):
            opd = weekDataDF['Octets/Duration'][index]
            if(timeRange == finalSplitList['Time'][len(finalSplitList)-1]):
                prev = finalSplitList['Octets/Duration'][len(finalSplitList)-1]  
                opdFinal = (prev*sameIndex+opd)/(sameIndex+1)
                finalSplitList.at[(len(finalSplitList)-1),'Octets/Duration']=opdFinal
                sameIndex=sameIndex+1
                flag=0
            else:
                finalSplitList = finalSplitList.append({'Time':timeRange, 'Octets/Duration':opd}, ignore_index=True)
                flag=0
                sameIndex=1
            index=index+1
            
        else:
            if(epochtime2<epochtime):
                index=index+1
            else:
                if(flag):
                    finalSplitList = finalSplitList.append({'Time':timeRange, 'Octets/Duration':0}, ignore_index=True)
                i=i+1
                flag=1
                
    i=i+1
    if (i<len(epochListDF)):
        while (i<len(epochListDF)):
            epochtime = int(epochListDF['Epoch list'][i])
            startTime = time.strftime("%A -- %H:%M:%S --", time.localtime(epochtime))
            endtime = epochtime+10
            endTime = time.strftime("%H:%M:%S", (time.localtime(endtime)))
            timeRange = startTime + endTime
            finalSplitList = finalSplitList.append({'Time':timeRange, 'Octets/Duration':0}, ignore_index=True)
            i=i+1

                   
    return finalSplitList

   
def z_val(r_1a2a, r_1a2b, r_2a2b, N):

    rm2 = ((r_1a2a ** 2) + (r_1a2b ** 2)) / 2
    f = (1 - r_2a2b) / (2 * (1 - rm2))
    h = (1 - f * rm2) / (1 - rm2)

    z_1a2a = 0.5 * (math.log10((1 + r_1a2a)/(1 - r_1a2a)))
    z_1a2b = 0.5 * (math.log10((1 + r_1a2b)/(1 - r_1a2b)))

    z = (z_1a2a - z_1a2b) * ((N-3) ** 0.5) / (2 * (1 - r_2a2b) * h)

    return z


def p_val(z):
    
    p = 0.3275911
    a1 = 0.254829592
    a2 = -0.284496736
    a3 = 1.421413741
    a4 = -1.453152027
    a5 = 1.061405429

    sign = None
    if z < 0.01:
        sign = -1
    else:
        sign = 1

    x = abs(z) / (2 ** 0.5)
    t = 1 / (1 + p * x)
    erf = 1.0 - (((((a5 * t + a4) * t) + a3) * t + a2) * t + a1) * t * math.exp(-x * x)

    return 0.5 * (1 + sign * erf)

def week_gen(files, delN):
    
    flags = []
    
    for i in range (0, len(files)):
        flags.append(0)

    firstweek = pd.DataFrame()
    secondweek = pd.DataFrame()
    epochListDF = pd.DataFrame()
    weeklyFinalDF = pd.DataFrame()
    
    index2=0
    for i in range(0, len(files)):
        
        if (flags[i]==0):
            f = files[i]
            filename = os.path.splitext(os.path.basename(f))[0]
            userDF = pd.read_excel(f)
            print(filename)
            
            firstweek, secondweek = getWeek(userDF)
            print(f)
            
            epochListDF = epochTimeCreate(firstweek)
            weeklyFinalDF = splitData(epochListDF, firstweek, delN)
            append_df_to_excel('results/'+str(delN)+'/'+filename+'_week_1.xlsx', weeklyFinalDF)
            print('results/'+filename+'_week_1.xlsx is done')
            
            epochListDF = epochTimeCreate(secondweek)
            weeklyFinalDF = splitData(epochListDF, secondweek, delN)
            append_df_to_excel('results/'+str(delN)+'/'+filename+'_week_2.xlsx', weeklyFinalDF)
            print('results/'+filename+'_week_2.xlsx is done')
            flags[i] = 1
    

def spearman_calc(files, tf):
    
    n = len(files)
    pTable = [[0 for x in range (n)] for y in range (n)]

    for i in range(0, len(files)-1):
        for j in range(i+1, len(files)):
           
            file1 = files[i]
            file2 = files[j]
            
            f1 =  os.path.splitext(os.path.basename(file1))[0]
            f2 =  os.path.splitext(os.path.basename(file2))[0]
            
            a1 = pd.read_excel('results/'+str(tf)+'/'+f1+'_week_1.xlsx')['Octets/Duration']
            a2 = pd.read_excel('results/'+str(tf)+'/'+f1+'_week_2.xlsx')['Octets/Duration']
            b1 = pd.read_excel('results/'+str(tf)+'/'+f2+'_week_1.xlsx')['Octets/Duration']
            b2 = pd.read_excel('results/'+str(tf)+'/'+f2+'_week_2.xlsx')['Octets/Duration']
            
            r_1a2a = scipy.stats.spearmanr(a1, a2)[0]
            r_1a2b = scipy.stats.spearmanr(a1, b2)[0]
            r_2a2b = scipy.stats.spearmanr(a2, b2)[0]
            r_1b2b = scipy.stats.spearmanr(b1, b2)[0]
            r_1b2a = scipy.stats.spearmanr(b1, a2)[0]
            
            z_value = z_val(r_1a2a, r_1a2b, r_2a2b, len(a1))
            print(z_value)
            p_value = p_val(z_value)
            print(p_value)
            pTable[i][j] = p_value
            
            z_value = z_val(r_1b2b, r_1b2a, r_2a2b, len(a1))
            print(z_value)
            p_value = p_val(z_value)
            print(p_value)
            final[j][i] = p_value
    
    finalDF = pd.DataFrame(pTable)
    append_df_to_excel('P_table_'+str(227)+'.xlsx', finalDF)

if __name__ == '__main__':
    
    #Loading all files
    files = glob.glob('data/*.xlsx')
    
    #Generating different time windows
    week_gen(files, 10)
    week_gen(files, 227)
    week_gen(files, 300)
    
    #Spearman's coeffient calculation and P-value generation
    n = len(files)
    
    spearman_calc(files, 10)
    spearman_calc(files, 227)
    spearman_calc(files, 300)
    


    
    
































